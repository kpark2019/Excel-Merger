# Author: Kenneth Park


import shutil
import os

slash = "\\"

import openpyxl
from openpyxl import *

# PLEASE EDIT BEFORE RUNNING
header_row_number = 4                                           # CHANGE DEPENDING ON HEADER ROW NUMBER
start_copy_col = 3
end_copy_col = 21

# DO NOT EDIT
full_target_file_name = input("파일 패치를 입력하십시오:\n")
path_name = os.getcwd()
split_name = full_target_file_name.split(slash)
target_file_name = split_name[len(split_name) - 1]      # get the target file name by getting the last in the list
region = target_file_name[:2]                               # get the region from the name of the file
copied_file_name = 'CopiedWorkbook.xlsx'
headers = ["지역","ro일자", "문서번호", "작업유형", "입금형태", "비고", "청구고객명", "판매고객명", "R/O번호", "보험구분",
           "보험접수번호", "차종", "차량번호", "판매딜러", "부품매출", "공임매출", "기타매출", "공급가액", "부가세액", "총금액"]

def make_wb_copy():
    """
    This method copies an existing Excel file and creates a duplicate
    to preserve the quality of the original.
    """
    shutil.copy(full_target_file_name, path_name + copied_file_name)     # copy the file


def paste_row(row, ws, current_row):
    """
    This method parses through each row of the original Excel file
    and either deletes the row if it is empty, or copies and pastes
    the row onto a merged Excel file after taking into account the
    last used row in the Excel file.
    """
    count_col = 1
    ws.cell(row=current_row, column=count_col).value = region
    count_col += 1
    for current_column in range(start_copy_col, end_copy_col + 1):                                  # parse through each selected column in the selected row
        ws.cell(row=current_row, column=count_col).value = row[current_column - 1].value       # paste the selected cell into the worksheet's current row
        count_col += 1


def parse_rows(wb, merged_ws):
    """
    This method removes empty rows from the
    Excel file in order to help cohesively merge
    the different sheets.
    """
    number_of_sheets = len(wb.worksheets) - 9                                   # Number of sheets
    for sheet_number in range(1, 1 + number_of_sheets):                         # parse through each sheet
        sheet = wb["%s%s" % (sheet_number, "월")]                               # determine sheet by number and Korean word for month
        for row_number in range(header_row_number + 1, sheet.max_row + 1):      # parse through each row (using the row number)
            row = sheet[row_number]                                             # set 'row_number's corresponding row to the row variable
                                                                                # start checking if it is an empty row.
            if row[3].value is None:                                            # if the cell is empty, move towards deleting the row
                    if row[14].value is None:                                   # double check with another cell
                        sheet.delete_rows(row_number, 1)                        # if both checkpoints are empty, delete that row
            else:                                                               # otherwise, if the cell is not empty, move towards copying and pasting
                current_row = merged_ws.max_row + 1                             # find the first empty row
                paste_row(row, merged_ws, current_row)                          # paste the selected row into the merged Excel file at the current row
                print(row_number)

def set_headers(ws):
    """
    This method adds headers to an existing
    Excel file.
    """
    for column in range(1, 1 + len(headers)):                       # parse through each column in the first row
        ws.cell(row=1, column=column).value = headers[column - 1]   # add corresponding header value to the Excel file


def create_new_workbook():
    """
    This method creates a new workbook to add
    the merged data.
    """
    merged_wb = openpyxl.Workbook()             # create new Workbook
    merged_wb["Sheet"].title = "Merged Data"    # change the title of the new sheet to "Merged Data"
    set_headers(merged_wb["Merged Data"])       # set the headers
    merged_wb.save('Merged_Data.xlsx')          # save the Workbook with the filename "Merged_Data.xlsx"
    return merged_wb


""" MAIN FUNCTIONS """


def main():

    # Open or create the Merged_Data.xlsx file.
    file_name = path_name + 'Merged_Data.xlsx'
    print("Opening Merged_Data.xlsx . . .")
    try:
        merged_wb = openpyxl.load_workbook(file_name)
        print("Opened Merged_Data.xlsx.")
    except IOError:
        print("IOError... Creating new Workbook")
        merged_wb = create_new_workbook()

    # Copy the target Excel file to an editable version.
    print("Making copy of target file . . .")
    make_wb_copy()
    print("Copied target file.")
    separated_wb = load_workbook(path_name + copied_file_name)

    # Parse through the rows and:
    #   1) Remove empty rows.
    #   2) Copy/Paste data row-by-row into merged Excel file.
    print("Parsing through rows . . .")
    parse_rows(separated_wb, merged_wb["Merged Data"])
    print("Parsed rows.")

    print("Saving files . . .")
    separated_wb.save("CopiedWorkbook.xlsx")
    separated_wb.close()
    merged_wb.save("Merged_Data.xlsx")
    merged_wb.close()
    print("Saved files.")

if __name__ == '__main__':
    main()